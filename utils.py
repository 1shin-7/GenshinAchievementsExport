import cv2
import keyboard
import numpy as np
import pyautogui
from paddleocr import PaddleOCR
import win32com.client
import time
import openpyxl


def minDistance(w1,w2):
    '''
    字符串最小编辑距离
    '''
    m,n = len(w1),len(w2)
    if(m==0):
	    return m
    if(n==0):
	    return n
    step = [[0]*(n+1)for _ in range(m+1)]
    for i in range(1,m+1):step[i][0]=i
    for j in range(1,n+1):step[0][j]=j
    for i in range(1,m+1):
	    for j in range(1,n+1):
		    if w1[i-1] == w2[j-1] :
			    diff=0
		    else:diff=1
		    step[i][j] = min(step[i-1][j-1],min(step[i-1][j],step[i][j-1]))+diff	
    return step[m][n]


def Search(a,list):
    '''
    在列表中模糊查找值
    '''
    for x in list:
        if minDistance(a,x)<3:
            return True
    return False


with open('已知栏目.txt','r') as f:
    txt=f.readlines()
LIST_names=[x.strip() for x in txt]
def Find(a):
    '''
    在已知栏目中匹配值
    '''
    dis=[]
    for name in LIST_names:
        dis.append(minDistance(a,name))
    return LIST_names[dis.index(min(dis))]


def cross(bbox,frameworks):
    '''
    判断矩形与一系列矩形是否相交
    '''
    for i,fw in enumerate(frameworks):
        center_fw=(fw[0]+fw[2]//2,fw[1]+fw[3]//2)
        center_bbox=(bbox[0]+bbox[2]//2,bbox[1]+bbox[3]//2)
        if abs(center_fw[0]-center_bbox[0])<(fw[2]+bbox[2])//2 and abs(center_fw[1]-center_bbox[1])<(fw[3]+bbox[3])//2:
            return i
    return -1

def get_rects(image,mode='left'):
    '''
    获得图片中左边或右边的一系列矩形
    '''
    image=cv2.cvtColor(image,cv2.COLOR_BGR2GRAY)
    thresh_min,thresh_max=(50,100)
    image=cv2.Canny(image,thresh_min,thresh_max)
    contours=cv2.findContours(image,cv2.CHAIN_APPROX_SIMPLE,cv2.RETR_CCOMP)[0]
    frameworks=[]
    if mode=='left': 
        for contour in contours:
            bbox=cv2.boundingRect(contour)
            if 100<=bbox[3]<=120 and (580<=bbox[2]<=700):
                if cross(bbox,frameworks)!=-1:
                    continue
                frameworks.append(bbox)
    else:
        
        for contour in contours:
            bbox=cv2.boundingRect(contour)
            if 110<=bbox[3]<=150 and (1000<=bbox[2]<=1500):
                if cross(bbox,frameworks)!=-1:
                    continue
                frameworks.append(bbox)
    frameworks.sort(key=lambda x:x[1])
    return frameworks


def drag_to_top(mode='right'):
    '''
    控制鼠标进行翻页
    '''
    xpos=1200 if mode=='right' else 200
    pyautogui.moveTo(xpos,900)
    pyautogui.dragTo(xpos,200,duration=1,tween=pyautogui.easeOutQuad,button='left')
    pyautogui.click(xpos,600)
    pyautogui.click(10,10)
    time.sleep(0.5)


END_image=cv2.imread('analysis/2.png',cv2.IMREAD_GRAYSCALE)
def is_end_left(text):
    '''
    判断是否翻页结束
    '''
    if text==LIST_names[-1]:
        return True
    return False


def is_end_right(image1):
    '''
    判断是否翻页结束
    '''
    left=780
    right=820
    up=990
    down=1020
    return cv2.PSNR(image1[up:down,left:right],END_image[up:down,left:right])>40


def hundreds_in(name):
    '''
    判断需不需要处理第一项
    '''
    return name!='天地万象' and name!='心跳的记忆'


ocr= PaddleOCR(lang='ch')
def get_right_rect_infos(image,fw):
    '''
    获得右侧矩形框的文字信息
    '''
    left=fw[0]
    right=fw[0]+fw[2]
    up=fw[1]
    down=fw[1]+fw[3]
    width=right-left 
    height=down-up 
    image_taskname=image[up:down-height*5//10,left+width*1//10:right-width*3//10]
    image_taskinfo=image[up+height*5//10:down,left+width*1//10:right-width*3//10]
    image_complete=image[up:down,left+width*85//100:right]
    taskname=ocr.ocr(image_taskname)[0][1][0]
    taskinfo=ocr.ocr(image_taskinfo)[0][1][0]
    complete=ocr.ocr(image_complete)[0][1][0]
    return taskname, taskinfo,complete


def get_left_rect_info(image,fw):
    '''
    获得左侧矩形框的文字信息
    '''
    left=fw[0]
    right=fw[0]+fw[2]
    up=fw[1]
    down=fw[1]+fw[3]
    width=right-left 
    height=down-up 
    result=ocr.ocr(image[up:down-height*5//10,left+width*1//10:right-width*2//10])
    text=result[0][1][0]
    text=Find(text)
    return text

def D_Click(fw):
    left,up,width,height=fw
    pyautogui.doubleClick(left+width//2,up+height//2)
    time.sleep(1)

def is_chinese(uchar):
    if uchar >= u'\u4e00' and uchar<=u'\u9fa5':
        return True
    else:
        return False


def is_number(uchar):
    if uchar >= u'\u0030' and uchar<=u'\u0039':
        return True
    else:
        return False

 
def is_alphabet(uchar):
    if (uchar >= u'\u0041' and uchar<=u'\u005a') or (uchar >= u'\u0061' and uchar<=u'\u007a'):
        return True
    else:
        return False


def is_other(uchar):
    '''
    是不是特殊符号
    '''
    if not (is_chinese(uchar) or is_number(uchar) or is_alphabet(uchar)):
        return True
    else:
        return False


def CH_EN(txt):
    '''
    过滤特殊符号
    '''
    ans=''
    for x in txt:
        if not is_other(x):
            ans+=x 
    return ans


def Search2(a,list):
    '''
    获得列表中值的索引及值的模糊匹配
    '''
    for i,x in enumerate(list):
        if minDistance(CH_EN(a),CH_EN(x))<3:
            return i,x
    return -1,-1


def compare_xlsx():
    '''
    比对表格
    '''
    workbook_read = openpyxl.load_workbook('test.xlsx')
    worksheet_read = workbook_read['Full']
    max_row=worksheet_read.max_row

    my_achievements=[]
    my_situation=[]
    for i in range(1,max_row):
        my_achievements.append(worksheet_read['c'+str(i+1)].value)
        my_situation.append('达成' if not '/' in worksheet_read['e'+str(i+1)].value else worksheet_read['e'+str(i+1)].value)
    workbook_read = openpyxl.load_workbook('2.4.0成就_CHS.xlsx')
    worksheet_read = workbook_read['Full']
    if  '未完成' in workbook_read.sheetnames:
        del workbook_read['未完成']
    worksheet_write=workbook_read.create_sheet('未完成')
    index_write=1
    max_row=worksheet_read.max_row

    for i in range(1,max_row):
        achievement=worksheet_read['b'+str(i+1)].value
        index,ans=Search2(achievement, my_achievements)
        if index==-1 or (index!=-1 and my_situation[index]!='达成'):
            worksheet_write['a'+str(index_write)].value=achievement
            worksheet_write['b'+str(index_write)].value=worksheet_read['c'+str(i+1)].value
            worksheet_write['c'+str(index_write)].value=ans
            worksheet_write['d'+str(index_write)].value='' if index==-1 else my_situation[index]
            index_write+=1
    for x in workbook_read.sheetnames:
        if x!='未完成':
            del workbook_read[x]
    workbook_read.save('compare_ans.xlsx')
def export_xlsx(Need_Speaker=False):
    '''
    导出表格
    '''
    def get_page(name):
        global worksheet_index
        if Need_Speaker:
            speak.Speak('开始整理'+name)
        text_list=[]
        info_list=[]
        complete_list=[]
        RUN=True
        while RUN:
            pyautogui.moveTo(10,10)
            image=pyautogui.screenshot()
            image=cv2.cvtColor(np.asarray(image),cv2.COLOR_RGB2BGR)
            frameworks=get_rects(image,'right')
            for i,fw in enumerate(frameworks):
                if i==0 and hundreds_in(name):
                    continue             
                taskname, taskinfo, complete = get_right_rect_infos(image,fw)                   
                if not Search(taskname,text_list):
                    text_list.append(taskname)
                    info_list.append(taskinfo)
                    complete_list.append(complete)
            if is_end_right(cv2.cvtColor(image,cv2.cv2.COLOR_BGR2GRAY)):
                if Need_Speaker:
                    speak.Speak(name+'整理完毕')
                for i,x in enumerate(text_list):                
                    worksheet_write.cell(worksheet_index,1).value=name
                    worksheet_write.cell(worksheet_index,2).value=i+1
                    worksheet_write.cell(worksheet_index,3).value=x 
                    worksheet_write.cell(worksheet_index,4).value=info_list[i]
                    if len(complete_list[i].split('/'))>1:
                        ans=complete_list[i]
                    else:
                        ans='达成'
                    worksheet_write.cell(worksheet_index,5).value=ans
                    worksheet_index+=1
                break
            drag_to_top(mode='right')
        
            

    def get_indexs():
        if Need_Speaker:
            speak.Speak('查找列表')
        text_list=[]    
        while True :        
            pyautogui.moveTo(10,10)
            image=pyautogui.screenshot()
            image=cv2.cvtColor(np.asarray(image),cv2.COLOR_RGB2BGR)
            frameworks=get_rects(image,'left') 
            for fw in frameworks:
                text= get_left_rect_info(image,fw)
                if not text in text_list:
                    text_list.append(text)
                    D_Click(fw)
                    get_page(text)
                    workbook_write.save('test.xlsx')
            if is_end_left(text) :
                if Need_Speaker:
                    speak.Speak('查找完毕') 
                break
            else:
                drag_to_top(mode='left')
    speak = win32com.client.Dispatch('SAPI.SPVOICE')
    workbook_write=openpyxl.Workbook()
    if 'Full' in workbook_write.sheetnames:
        del workbook_write['Full']
    worksheet_write=workbook_write.create_sheet('Full')    
    worksheet_write.cell(1,1).value='所属栏目'
    worksheet_write.cell(1,2).value='序号'
    worksheet_write.cell(1,3).value='名称'
    worksheet_write.cell(1,4).value='内容'
    worksheet_write.cell(1,5).value='达成情况'
    global worksheet_index
    worksheet_index=2
    keyboard.wait('r')
    if Need_Speaker:
        speak.Speak('程序启动') 
    get_indexs()

if __name__=='__main__':
    keyboard.wait('r')
    
        
    