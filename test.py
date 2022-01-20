
import openpyxl
from main import minDistance
def is_chinese(uchar):
        if uchar >= u'\u4e00' and uchar<=u'\u9fa5':

                return True

        else:

                return False

 

def is_number(uchar):
        if uchar >= u'\u0030' and uchar<=u'\u0039':

                return True

        else:

                return False

 

def is_alphabet(uchar):
        if (uchar >= u'\u0041' and uchar<=u'\u005a') or (uchar >= u'\u0061' and uchar<=u'\u007a'):

                return True

        else:

                return False

 

def is_other(uchar):
        if not (is_chinese(uchar) or is_number(uchar) or is_alphabet(uchar)):

                return True

        else:

                return False

def CH_EN(txt):
    ans=''
    for x in txt:
        if not is_other(x):
            ans+=x 
    return ans
def Search(a,list):
    for i,x in enumerate(list):
        if minDistance(CH_EN(a),CH_EN(x))<3:
            return i,x
    return -1,-1

workbook_read = openpyxl.load_workbook('test.xlsx')
worksheet_read = workbook_read['Full']
max_row=worksheet_read.max_row

my_achievements=[]
my_situation=[]
for i in range(1,max_row):
    my_achievements.append(worksheet_read['c'+str(i+1)].value)
    my_situation.append('达成' if not '/' in worksheet_read['e'+str(i+1)].value else worksheet_read['e'+str(i+1)].value)
workbook_read = openpyxl.load_workbook('2.4.0成就_CHS.xlsx')
worksheet_read = workbook_read['Full']
if  '未完成' in workbook_read.sheetnames:
    del workbook_read['未完成']
worksheet_write=workbook_read.create_sheet('未完成')
index_write=1
max_row=worksheet_read.max_row

for i in range(1,max_row):
    achievement=worksheet_read['b'+str(i+1)].value
    index,ans=Search(achievement, my_achievements)
    if index==-1 or (index!=-1 and my_situation[index]!='达成'):
        worksheet_write['a'+str(index_write)].value=achievement
        worksheet_write['b'+str(index_write)].value=worksheet_read['c'+str(i+1)].value
        worksheet_write['c'+str(index_write)].value=ans
        worksheet_write['d'+str(index_write)].value='' if index==-1 else my_situation[index]
        index_write+=1
for x in workbook_read.sheetnames:
    if x!='未完成':
        del workbook_read[x]
workbook_read.save('compare_ans.xlsx')

