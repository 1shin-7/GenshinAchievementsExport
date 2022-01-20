import cv2
import keyboard
import numpy as np
import pyautogui
from paddleocr import PaddleOCR
import win32com.client
import time
import openpyxl


def minDistance(w1,w2):
	m,n = len(w1),len(w2)
	if(m==0):
		return m
	if(n==0):
		return n
	step = [[0]*(n+1)for _ in range(m+1)]
	for i in range(1,m+1):step[i][0]=i
	for j in range(1,n+1):step[0][j]=j
	for i in range(1,m+1):
		for j in range(1,n+1):
			if w1[i-1] == w2[j-1] :
				diff=0
			else:diff=1
			step[i][j] = min(step[i-1][j-1],min(step[i-1][j],step[i][j-1]))+diff	
	return step[m][n]
def Search(a,list):
    for x in list:
        if minDistance(a,x)<3:
            return True
    return False

def Find(a):
    with open('已知栏目.txt','r') as f:
        txt=f.readlines()
    names=[x.strip() for x in txt]
    dis=[]
    for name in names:
        dis.append(minDistance(a,name))
    return names[dis.index(min(dis))]
def cross(bbox,frameworks):
    for i,fw in enumerate(frameworks):
        center_fw=(fw[0]+fw[2]//2,fw[1]+fw[3]//2)
        center_bbox=(bbox[0]+bbox[2]//2,bbox[1]+bbox[3]//2)
        if abs(center_fw[0]-center_bbox[0])<(fw[2]+bbox[2])//2 and abs(center_fw[1]-center_bbox[1])<(fw[3]+bbox[3])//2:
            return i
    return -1

def get_rects(image,mode='left'):
    image=cv2.cvtColor(image,cv2.COLOR_BGR2GRAY)
    thresh_min,thresh_max=(50,100)
    image=cv2.Canny(image,thresh_min,thresh_max)
    contours,hierachy=cv2.findContours(image,cv2.CHAIN_APPROX_SIMPLE,cv2.RETR_CCOMP)
    frameworks=[]
    if mode=='left': 
        for contour in contours:
            bbox=cv2.boundingRect(contour)
            if 100<=bbox[3]<=120 and (580<=bbox[2]<=700):
                if cross(bbox,frameworks)!=-1:
                    continue
                frameworks.append(bbox)
    else:
        
        for contour in contours:
            bbox=cv2.boundingRect(contour)
            if 110<=bbox[3]<=150 and (1000<=bbox[2]<=1500):
                if cross(bbox,frameworks)!=-1:
                    continue
                frameworks.append(bbox)
    frameworks.sort(key=lambda x:x[1])
    return frameworks

def get_page(name):
    global worksheet_index
    if Need_Speaker:
        speak.Speak('开始整理'+name)
    text_list=[]
    info_list=[]
    complete_list=[]
    RUN=True
    while RUN:
        pyautogui.moveTo(10,10)
        image=pyautogui.screenshot()
        image=cv2.cvtColor(np.asarray(image),cv2.COLOR_RGB2BGR)
        frameworks=get_rects(image,'right')
        all_seen=True 
        for i,fw in enumerate(frameworks):
            if i==0:
                if name!='天地万象' and name!='心跳的记忆':
                    continue 
            left=fw[0]
            right=fw[0]+fw[2]
            up=fw[1]
            down=fw[1]+fw[3]
            width=right-left 
            height=down-up 
            image_taskname=image[up:down-height*5//10,left+width*1//10:right-width*3//10]
            image_taskinfo=image[up+height*5//10:down,left+width*1//10:right-width*3//10]
            image_complete=image[up:down,left+width*85//100:right]
            cv2.rectangle(image,(left+width*1//10,up),(right-width*3//10,down-height*5//10),(255,0,0),3)
            cv2.rectangle(image,(left+width*1//10,up+height*5//10),(right-width*3//10,down),(0,255,0),3)
            cv2.rectangle(image,(left+width*85//100,up+height*1//10),(right,down),(0,0,255),3)
            taskname=ocr.ocr(image_taskname)[0][1][0]
            taskinfo=ocr.ocr(image_taskinfo)[0][1][0]
            complete=ocr.ocr(image_complete)[0][1][0]
            if not Search(taskname,text_list):
                text_list.append(taskname)
                info_list.append(taskinfo)
                complete_list.append(complete)
                all_seen=False
        if all_seen:
            if Need_Speaker:
                speak.Speak(name+'整理完毕')
            for i,x in enumerate(text_list):                
                worksheet_write.cell(worksheet_index,1).value=name
                worksheet_write.cell(worksheet_index,2).value=i+1
                worksheet_write.cell(worksheet_index,3).value=x 
                worksheet_write.cell(worksheet_index,4).value=info_list[i]
                if len(complete_list[i].split('/'))>1:
                    ans=complete_list[i]
                else:
                    ans='达成'
                worksheet_write.cell(worksheet_index,5).value=ans
                worksheet_index+=1
            break
        lastbox=frameworks[-1]
        xpos=lastbox[0]+lastbox[2]//2+100
        ypos=lastbox[1]+lastbox[3]//2
        pyautogui.moveTo(xpos,ypos)
        pyautogui.dragTo(xpos,200,duration=1,tween=pyautogui.easeOutQuad,button='left')
        pyautogui.click(xpos,200)
        
            

def get_indexs():
    if Need_Speaker:
        speak.Speak('查找列表')
    text_list=[]
    RUN=True
    while RUN:        
        pyautogui.moveTo(10,10)
        image=pyautogui.screenshot()
        image=cv2.cvtColor(np.asarray(image),cv2.COLOR_RGB2BGR)
        frameworks=get_rects(image,'left')
        all_seen=True 
        for fw in frameworks:
            left=fw[0]
            right=fw[0]+fw[2]
            up=fw[1]
            down=fw[1]+fw[3]
            width=right-left 
            height=down-up 
            result=ocr.ocr(image[up:down-height*5//10,left+width*1//10:right-width*2//10])
            text=result[0][1][0]
            text=Find(text)
            if not text in text_list:
                text_list.append(text)
                pyautogui.doubleClick(left+width//2,up+height//2)
                time.sleep(1)
                get_page(text)
                workbook_write.save('test.xlsx')
                all_seen=False
        if all_seen:
            if Need_Speaker:
                speak.Speak('查找完毕') 
            break
        else:
            lastbox=frameworks[-1]
            xpos=lastbox[0]+lastbox[2]//2+100
            ypos=lastbox[1]+lastbox[3]//2
            pyautogui.moveTo(xpos,ypos)
            pyautogui.dragTo(xpos,10,duration=1.5,tween=pyautogui.easeOutQuad,button='left')
            time.sleep(1)
if __name__=='__main__':
    speak = win32com.client.Dispatch('SAPI.SPVOICE')
    ocr= PaddleOCR(lang='ch')
    workbook_write=openpyxl.Workbook()
    if 'Full' in workbook_write.sheetnames:
        del workbook_write['Full']
    worksheet_write=workbook_write.create_sheet('Full')    
    worksheet_write.cell(1,1).value='所属栏目'
    worksheet_write.cell(1,2).value='序号'
    worksheet_write.cell(1,3).value='名称'
    worksheet_write.cell(1,4).value='内容'
    worksheet_write.cell(1,5).value='达成情况'

    worksheet_index=2
    keyboard.wait('r')

    Need_Speaker=False

    if Need_Speaker:
        speak.Speak('程序启动') 
    get_indexs()

